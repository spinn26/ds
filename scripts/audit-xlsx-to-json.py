"""
Stage 1 of the audit-driven products import pipeline.

Reads a Google Sheets export (XLSX) of «Аудит Продукты и баллы без учёта НДС»,
parses the «Продукты» sheet, deduplicates rate lines down to unique programs
(year is a parameter of the program, not a separate program — so EVO 25/20/15/10
collapse into a single EVO program with N rate lines), captures the fill-color
decision (🔴 red = deactivate, anything else = keep active), and emits a JSON
file the PHP importer reads.

Usage:
    python scripts/audit-xlsx-to-json.py <input.xlsx> <output.json>

Defaults if args omitted:
    input:  Db/audit_products_2026_05_22.xlsx
    output: storage/app/audit-programs.json
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

RED = "FFFF0000"
GRAY = "FFB7B7B7"
SKIP_COLORS = {None, "00000000", "FFFFFFFF", "theme:0"}

# Columns on the «Продукты» sheet (1-based indices align with openpyxl row tuples — 0-based here).
# A row layout: B=комментарий, C=ТИП, D=ПРОДУКТ, E=ПРОГРАММА, F=Стоимость, G=ВАЛЮТА,
#               H=ПОСТАВЩИК, I=% DS, J=Свойство, K=Срок Контракта, L=Год выплаты КВ,
#               M=Баллы, N=МЕТОДИКА, O=Комментарии, P=Категория
COL = {
    "comment": 1,    # B
    "type": 2,       # C
    "product": 3,    # D
    "program": 4,    # E
    "cost": 5,       # F
    "currency": 6,   # G
    "vendor": 7,     # H
    "ds_percent": 8, # I
    "property": 9,   # J
    "term": 10,      # K
    "year": 11,      # L
    "points": 12,    # M
    "formula": 13,   # N
    "extra": 14,     # O
    "category": 15,  # P
}


def cell_color(cell) -> str | None:
    f = cell.fill
    if not f or not f.fgColor or f.patternType != "solid":
        return None
    fg = f.fgColor
    if fg.type == "rgb":
        return fg.rgb
    if fg.type == "theme":
        return f"theme:{fg.theme}"
    return None


def cell_value(cell) -> str | None:
    if cell.value is None:
        return None
    s = str(cell.value).strip()
    return s or None


def main(in_path: Path, out_path: Path) -> int:
    if not in_path.exists():
        print(f"input not found: {in_path}", file=sys.stderr)
        return 1

    wb = load_workbook(in_path, data_only=True)
    sheet = next((s for s in wb.sheetnames if "родукт" in s), wb.sheetnames[0])
    ws = wb[sheet]

    programs: dict[tuple[str, str, str], dict[str, Any]] = {}

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column),
        start=1,
    ):
        if len(row) < 16:
            continue
        typ = cell_value(row[COL["type"]])
        product = cell_value(row[COL["product"]])
        program = cell_value(row[COL["program"]])
        if not typ or not program or not product:
            continue
        # heuristic: a long sentence in "type" column is a section banner or commentary, skip
        if len(typ) > 60:
            continue
        # column-index header row («3 | 4 | 5 …») — skip if all three fields are bare ints
        if typ.isdigit() and product.isdigit() and program.isdigit():
            continue
        # the «Тарифы для продуктов Инсмарт» reference link row
        if program.startswith("http"):
            continue

        # Row-level dominant non-gray non-skip color = audit decision for this rate line
        row_colors: Counter[str] = Counter()
        for c in row:
            col = cell_color(c)
            if col not in SKIP_COLORS and col != GRAY:
                row_colors[col] += 1

        key = (typ, product, program)
        rec = programs.setdefault(
            key,
            {
                "type": typ,
                "product": product,
                "program": program,
                "vendors": set(),
                "currencies": set(),
                "categories": set(),
                "years": set(),
                "terms": set(),
                "rate_lines": 0,
                "row_colors": Counter(),
                "first_row": r_idx,
                "last_row": r_idx,
                "comment_snippets": set(),
                "tariffs": [],
            },
        )
        rec["rate_lines"] += 1
        rec["last_row"] = r_idx
        rec["row_colors"].update(row_colors)

        vendor   = cell_value(row[COL["vendor"]])
        currency = cell_value(row[COL["currency"]])
        category = cell_value(row[COL["category"]])
        year     = cell_value(row[COL["year"]])
        term     = cell_value(row[COL["term"]])
        ds_pct   = cell_value(row[COL["ds_percent"]])
        prop     = cell_value(row[COL["property"]])
        cost     = cell_value(row[COL["cost"]])
        points   = cell_value(row[COL["points"]])
        formula  = cell_value(row[COL["formula"]])
        comment  = cell_value(row[COL["comment"]])
        extra    = cell_value(row[COL["extra"]])

        if vendor:   rec["vendors"].add(vendor)
        if currency: rec["currencies"].add(currency)
        if category: rec["categories"].add(category)
        if year:     rec["years"].add(year)
        if term:     rec["terms"].add(term)
        if comment and len(comment) < 80:
            rec["comment_snippets"].add(comment)

        # Per-tariff-line capture: this is where year-as-parameter lives.
        # One audit row = one tariff row in the JSONB column on program_catalog.
        rec["tariffs"].append({
            "row":       r_idx,
            "vendor":    vendor,
            "currency":  currency,
            "ds_percent": ds_pct,
            "property":  prop,
            "term":      term,
            "year":      year,
            "fixed_cost": cost,
            "points":    points,
            "formula":   formula,
            "comment":   extra,
        })

    out_programs = []
    for rec in programs.values():
        colors = rec["row_colors"]
        dominant = colors.most_common(1)[0][0] if colors else None
        out_programs.append(
            {
                "type": rec["type"],
                "product": rec["product"],
                "program": rec["program"],
                "vendors": sorted(rec["vendors"]),
                "currencies": sorted(rec["currencies"]),
                "categories": sorted(rec["categories"]),
                "years": sorted(rec["years"]),
                "terms": sorted(rec["terms"]),
                "comment_snippets": sorted(rec["comment_snippets"]),
                "rate_lines": rec["rate_lines"],
                "has_red": colors.get(RED, 0) > 0,
                "dominant_color": dominant,
                "row_colors": dict(colors),
                "first_row": rec["first_row"],
                "last_row": rec["last_row"],
                "tariffs": rec["tariffs"],
            }
        )

    out_programs.sort(key=lambda r: (r["type"] or "", r["product"] or "", r["program"] or ""))

    payload = {
        "source": str(in_path.name),
        "sheet": sheet,
        "extracted_at_rows": ws.max_row,
        "summary": {
            "unique_programs": len(out_programs),
            "to_deactivate_red": sum(1 for p in out_programs if p["has_red"]),
            "to_keep_active": sum(1 for p in out_programs if not p["has_red"]),
        },
        "programs": out_programs,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    s = payload["summary"]
    print(f"wrote {out_path}")
    print(f"  unique programs: {s['unique_programs']}")
    print(f"  red (deactivate): {s['to_deactivate_red']}")
    print(f"  keep active:     {s['to_keep_active']}")
    return 0


if __name__ == "__main__":
    args = sys.argv[1:]
    here = Path(__file__).resolve().parent.parent
    in_path = Path(args[0]) if len(args) >= 1 else here / "Db" / "audit_products_2026_05_22.xlsx"
    out_path = Path(args[1]) if len(args) >= 2 else here / "storage" / "app" / "audit-programs.json"
    sys.exit(main(in_path, out_path))
